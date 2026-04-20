from pathlib import Path

from .base import BaseParser, ParsedDocument, ParsedPage


def _sheet_to_markdown(ws) -> str:
    """Convert an openpyxl worksheet to a markdown table."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""

    md_rows = []
    for i, row in enumerate(rows):
        cells = [str(cell) if cell is not None else "" for cell in row]
        md_rows.append("| " + " | ".join(cells) + " |")
        if i == 0:
            md_rows.append("| " + " | ".join(["---"] * len(cells)) + " |")
    return "\n".join(md_rows)


class ExcelParser(BaseParser):
    def can_parse(self, file_path: Path) -> bool:
        return file_path.suffix.lower() in {".xlsx", ".xls", ".xlsm"}

    def parse(self, file_path: Path) -> ParsedDocument:
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required: pip install openpyxl")

        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        pages: list[ParsedPage] = []

        for i, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            md = _sheet_to_markdown(ws)
            pages.append(
                ParsedPage(
                    page_number=i,
                    text=f"Sheet: {sheet_name}",
                    tables=[md] if md else [],
                )
            )

        return ParsedDocument(
            document_id=file_path.stem,
            filename=file_path.name,
            file_format="xlsx",
            pages=pages,
            metadata={"sheet_count": len(wb.sheetnames), "sheet_names": wb.sheetnames},
        )
